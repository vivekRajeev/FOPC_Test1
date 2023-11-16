import time
import openpyxl
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException
from xlrd import sheet
from selenium.webdriver.common.by import By
from collections import Counter
from google.oauth2 import service_account
from googleapiclient.discovery import build

from PageObjects.all_makes import Makes
from Utility.BaseClass import Baseclass
from google.oauth2 import service_account
from googleapiclient.discovery import build

class TestMakes_validated(Baseclass):
    def test_validate_makes(self):
        self.initialization()
        manufacture = []
        make_list = []
        manufacture_excel = []
        make_excel = []
        all_data = []
        all_data1 = []
        univalues = []
        ex_manufacture = []
        ex_manufacture_filt = []
        ex_manufacture_filt_excel = []
        ex_make = []
        make = Makes(self.driver)
        make.reference().click()
        time.sleep(2)
        make.makes().click()
        time.sleep(3)
        manufactures = make.manufact()
        time.sleep(2)
        store_name = make.store_names()
        for validmanufacture in manufactures:
            manufacture.append(validmanufacture.text)
        print(manufacture)
        time.sleep(2)
        makes = make.makedatas()
        for validmake in makes:
            make_list.append(validmake.text)
            make1 = set(make_list)
        print(make1)
        time.sleep(3)
        make.makefilter().click()
        time.sleep(2)
        make_name = make.makename().text
        # item = "MAZDA"
        # print(make_name)
        # book = openpyxl.load_workbook("C:\\Users\\windows\\Videos\\All_Makes.xlsx")
        book = openpyxl.load_workbook("C:\\Users\\windows\\PycharmProjects\\Makes_Excel_Validation\\TestData\\All_Makes.xlsx")
        make_excel_names = book.sheetnames
        print(make_excel_names)

        for make_excel_name in make_excel_names:
            if make_excel_name == make_name:

                MAKE = book[make_name]
                for row in MAKE.iter_rows(min_row=1, values_only=True):
                    manufacture_excel.append(row[0])
                    make_excel.append(row[1])
                    make_excel1 = set(make_excel)
                print(manufacture_excel)
                print(make_excel1)
                if(manufacture == manufacture_excel):
                    print(f"{store_name.text} manufactures matches in both UI and excel")
                else:
                    print(f"{store_name.text} manufactures is not matches both UI and excel")
                time.sleep(2)
                if (make1 == make_excel1):
                    print(f"{store_name.text} makes matches in both UI and excel")
                else:
                    print(f"{store_name.text} makes is not matches in both UI and excel")
                break
            else:
                print("make not present")
        time.sleep(2)

        make.excludemake().click()
        time.sleep(2)
        make.makefilter().click()
        time.sleep(2)
        make.namesend().send_keys(make_name)
        time.sleep(2)
        try:
            include_make = make.makename()
        except NoSuchElementException:
            print(f"Include_make {make_name} is not in Excluded makes")
        time.sleep(2)
        manufacture_excel.clear()
        make_excel.clear()
        make_excel1.clear()
        manufacture.clear()
        make1.clear()
        make.reset_layout().click()
        time.sleep(2)
        make.makefilter().click()
        time.sleep(2)
        ag_grid = make.ag_grid()

        while len(all_data) < 36:
            try:
                rows = make.row_data()
                for i in range(0, len(rows)):
                    row_data = rows[i].text
                    all_data.append(row_data)

                scroll_script = "arguments[0].scrollTop += arguments[0].clientHeight"
                time.sleep(2)
                self.driver.execute_script(scroll_script, ag_grid)
                time.sleep(2)
            except StaleElementReferenceException:
                time.sleep(5)
                ag_grid = make.ag_grid()

        for row_data in all_data:
            # print(row_data)
            all_data1.append(row_data)
            time.sleep(2)
        unique_values = set(all_data1)
        univalues = list(unique_values)
        print(len(univalues))
        # self.driver.find_element(By.XPATH, "(//span[@class='MuiTab-wrapper'])[2]").click()
        # self.driver.find_element(By.XPATH, "(//span[@class='ag-icon ag-icon-filter'])[3]").click()
        # univaluess = ["GM","HONDA","CHRYSLER"]
        for univalue in univalues:
            time.sleep(3)
            make.namesend().send_keys(univalue)
            time.sleep(4)
            make.fil_sel().click()
            time.sleep(3)
            ex_mannu = make.ex_manufac()
            ex_mak = make.ex_makes()

            for ex_manufact in ex_mannu:
                ex_manufacture.append(ex_manufact.text)
                filtered_array = [value for value in ex_manufacture if value.strip() != ""]
            ex_manufacture_filt.append(filtered_array)
            print(ex_manufacture_filt)
            time.sleep(4)
            for ex_ma in ex_mak:
                ex_make.append(ex_ma.text)
            print(ex_make)
            time.sleep(4)
            # book = openpyxl.load_workbook("C:\\Users\\windows\\Videos\\All_Makes.xlsx")
            book = openpyxl.load_workbook("C:\\Users\\windows\\PycharmProjects\\Makes_Excel_Validation\\TestData\\All_Makes.xlsx")

            make_excel_names = book.sheetnames
            # print(make_excel_names)
            for make_excel_name in make_excel_names:
                if make_excel_name == univalue:
                    MAKE = book[univalue]
                    for row in MAKE.iter_rows(min_row=1, values_only=True):
                        manufacture_excel.append(row[0])
                        filtered_array_excel = [value for value in manufacture_excel if value.strip() != ""]
                        make_excel.append(row[1])
                    ex_manufacture_filt_excel.append(filtered_array_excel)
                    # make_excel1 = set(make_excel)
                    print(ex_manufacture_filt_excel)
                    print(make_excel)
                    time.sleep(3)
                    if(ex_manufacture_filt == ex_manufacture_filt_excel):
                        print(f"{store_name.text} manufactures matches in both UI and excel")
                    else:
                        print(f"{store_name.text} manufactures is not matches both UI and excel")
                    time.sleep(2)
                    if (ex_make == make_excel):
                        print(f"{store_name.text} makes matches in both UI and excel")
                    else:
                        print(f"{store_name.text} makes is not matches in both UI and excel")
                    break
                else:
                    print("sheet name not matches")

            make.reset_layout().click()
            time.sleep(2)
            make.makefilter().click()
            time.sleep(2)
            ex_manufacture.clear()
            filtered_array.clear()
            ex_manufacture_filt.clear()
            ex_make.clear()
            manufacture_excel.clear()
            filtered_array_excel.clear()
            ex_manufacture_filt_excel.clear()
            make_excel.clear()
